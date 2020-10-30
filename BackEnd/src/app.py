from sqls import sqls
from datetime import datetime, date
from tqdm.auto import tqdm
from mysql import connector as c
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.cell import Cell
from pyspin.spin import Spin1, Spinner
from utils import utils
import calendar
import time
import pandas as pds
import pyodbc as pd
import numpy as np
import sys
import locale
from builtins import FileNotFoundError
import logging

logging.basicConfig(level=logging.DEBUG)
#logging.basicConfig(filename='/var/log/flask/main.proc.log',level=logging.DEBUG)
class PMPG():
    def __init__(self):
        #self.url_vsc = 'Driver={SQL Server};server=189.85.23.20,25789;DATABASE=VSCDB;UID=everton.kopec;PWD=Nova@123'
        self.url_vsc = 'Driver={ODBC Driver 17 for SQL Server};server=189.85.23.20,25789;DATABASE=VSCDB;UID=everton.kopec;PWD=Nova@123'
        # self.url_nova = 'DRIVER={MySQL ODBC 8.0 ANSI Driver};SERVER=10.85.24.17;' \
        #                 'PORT=3306;DATABASE=NovaFibra;UID=pfa;PWD=NovaFibr@2020;charset=utf8mb4'
        self.url_nova = 'Driver={MySQL ODBC 8.0 ANSI Driver};server=10.85.24.33;PORT=3306;DATABASE=novafibra;UID=pfa;PWD=NovaFibr@2020'
        self.url_nova_proc = c.connect(user='pfa', password='NovaFibr@2020', db='novafibra', host='10.85.24.33', auth_plugin='mysql_native_password')
        self.sql = sqls()
        self.listGroupsCreated = []
        self.groups = {'PMPG': ['PMPG', 'PMPG_0800', 'SME_ESCOLA', 'SME_CMEI']
                    , 'FMS': ['FMS_AIH', 'FMS_AIH_0800', 'FMS_PAB', 'FMS_PAB_0800']}

    # Create Connection
    def createConn(self):
        conn_vsc = pd.connect(self.url_vsc)
        conn_nova = pd.connect(self.url_nova)
        conn_nova_proc = self.url_nova_proc
        return conn_vsc, conn_nova, conn_nova_proc
    

    def execOperation(self, exec, ins, *values):
        """
        Execute Operation
        INS - Instrução - 0 - Para Select na Base do Vsc
        INS - Instrução - 1 - Para Insert na base da NovaFibra
        INS - Instrução - 2 - Select na base da NovaFibra
        INS - Instrução - 3 - Create Groups Views from Procedures
        INS - Instrução - 4 - Execute Procedures
        INS - Instrução - n - Create,Drop,update,delete Table na Base NovaFibra
        O restante ainda está pra ser definido,
        porém temos operações de Create e Drop/Table
        """
        cur_vsc = self.createConn()[0].cursor()
        cur_nova = self.createConn()[1].cursor()
        cur_nova_proc = self.createConn()[2].cursor()
        # print("Exec %s instrução %s" % (exec, ins))
        if ins == 0:
            cur_vsc.execute(str(exec))
            # cur_vsc.close()
            return cur_vsc.fetchall()
        elif ins == 1:
            for data in values:
                logging.debug('data: %s ' % (len(data)))
                cur_nova.executemany(str(exec), data)

        elif ins == 2:
            cur_nova.execute(exec)
            data = cur_nova.fetchall()
            return data
        elif ins == 3:
            cur_nova_proc.callproc(exec[0], [exec[1],exec[2]])
            for group in cur_nova_proc.stored_results():
                data = pds.DataFrame(group.fetchall())
            
            return data
        elif ins == 4:
            cur_nova_proc.callproc(exec[0], [exec[1][0], exec[1][1]])

        elif ins == 5:
            cur_nova_proc.callproc(exec[0], [exec[1]])
        else:
            cur_nova.execute(exec)
        
        cur_nova.commit()
        cur_nova.close()
        cur_vsc.close()

    
    # Put Data into NovaFibra
    def putData(self, data, table, *month):
        if len(month) >= 0:
            tables = table
        # if len(month) > 0:
        #     table = table + '_' + month[0]
        ins = data
        num_columns = len(data[0])
        # print(num_columns)
        # A função abaixo passa o número de colunas dinamicamente
        # tirando a necessidade de inserir manualmente a quantidade
        insert = "REPLACE INTO {} VALUES ({})".format(tables, ','.join(['?' for column in range(num_columns)]))
        self.execOperation(insert, 1, ins)
        logging.debug("The values has been inserted on Table: %s!" % table)
    
    #Método Temporário
    # Insert Groups
    def insertGroups(self, month, *groups):
        listGroups = list(groups)

        for i in listGroups:
            logging.debug("\n Creating Table --> %s " % i)
            view = self.sql.insertGroupsProc(month, i)
            proc = view[0][0]
            mnt = view[0][1]
            gp = view[0][2][0]
            self.execOperation([proc, [mnt, gp]], view[1])
            self.listGroupsCreated.append(i)
            # print("\n The Table %s has been created \n\n" % i)

        return logging.debug(" The Tables %s have been inserted \n " % self.listGroupsCreated)

    # Import Data
    # From: VSC Database
    # To: Nova Fibra Database
    def importVSCtoNF(self, month):
        
        getData = self.sql.getDataFaturamento("faturamento", month)
        results = self.execOperation(getData[0], getData[1], getData[2])
        data = []
        for i in results:
            data.append(i)
        self.putData(data, getData[2])
        return logging.debug("Import Data executed successfully!")


    # time sum up and cost - Used for exportExcel
    def sumTimeCost(self, df):
        totalSecs = 0
        group = pds.DataFrame.from_records(df, columns=["Tipo", "Soma_Duracao", "Soma_Custo"])
        group["Soma_Duracao"] = group["Soma_Duracao"].str.split(':')
        # Adding time
        for i in group["Soma_Duracao"]:
            timeparts = [int(s) for s in i]
            totalSecs += (timeparts[0] * 60 + timeparts[1]) * 60
        totalSecs, sec = divmod(totalSecs, 60)
        min = divmod(totalSecs, 60)
        # Put time to total
        total_time = "%s min %s seg" % (min)
        # Adding cost total and put into Series
        # after add to DataFrame Group
        total_cost = group["Soma_Custo"].sum()
        df2 = pds.Series(['Total', total_time, total_cost], index=["Tipo", "Soma_Duracao", "Soma_Custo"])
        group = group.append(df2, ignore_index=True)
        # Put time format to:
        # 3405 min 36 sec
        for i in range(len(group["Soma_Duracao"]) - 1):
            group["Soma_Duracao"][i] = group["Soma_Duracao"][i][0] + ' min ' + group["Soma_Duracao"][i][1] + ' seg'

        return group
    
    # Formating Money for exportExcel
    def moneytoReal(self, cost):
        a = '{:,.2f}'.format(float(cost))
        b = a.replace(',', 'v')
        c = b.replace('.', ',')
        return c.replace('v', '.')

    # Write Ext Excel
    """ data - data to be processed 
        group - group to be written
        file_to_write - the file will be written
        position - position to be written """
    def writeExcelExt(self, data, group, file_to_write):
        if (len(data)) == 0:
            logging.debug('\n -:No Data at this position:-\n')
            return

        if group == 'PMPG':
            group = 'Faturar_PMPG_SME'
        else:
            group = 'Faturar_FMS'

        data = data
        df_d = pds.DataFrame.from_records(data, columns=['QTY','GROUPS','POS_COL', 'POS_ROW', 'MONTH'])
        df_data = df_d[['QTY']]
        pos_row_data = df_d['POS_ROW'][0]
        # month_id = {'month':[pds.to_datetime(month, format='%Y%m%d')]}
        # df_month = pds.DataFrame(data=month_id,columns=['month'])
        # print('Mês', df_d['MONTH']) 
        
        self.append_df_to_excel(file_to_write, df_data, sheet_name=group, startrow=pos_row_data, header=None, index=False, startcol=3)
        logging.debug('--->Groups: %s<--- \n--->Qty: %s<--- \n--->Pos ROW: %s<---\n ' % (df_d['GROUPS'][0], df_d['QTY'][0], pos_row_data))

    # Grava o Mês da Planilha
    def writeExcelMonth(self, data, group, file_to_write):
        logging.debug(' len -> data %s ' % len(data))
        if (len(data)) == 0:
            logging.debug('\n -:No Data at this position:-\n')
            return
        
        if group == 'PMPG':
            group = 'Utilizacao_PMPG_SME'    
        else:
            group = 'Utilizacao_FMS'
        logging.debug('group choosed %s ' % group)
        month_id = {'month':[pds.to_datetime(data, format='%Y%m%d')]}
        df = pds.DataFrame(data=month_id, columns=['month'])
        self.append_df_to_excel(file_to_write, df, sheet_name=group, startrow=4, header=None, index=False, startcol=4)

  
    # Write Prop Ext Values
    def writeExcelProp(self, prop, group, file_to_write):
        if (len(prop)) == 0:
            logging.debug('\n -:No Proportional at this position:-\n')
            return

        if group == 'PMPG':
            group = 'Faturar_PMPG_SME'
        else:
            group = 'Faturar_FMS'

        # print('prop %s %s ' % (prop[0][0], prop[0][1]))
        prop = prop
        df_p = pds.DataFrame.from_records(prop, columns=['GROUPS','PROP','POS_COL','POS_ROW'])
        df_prop = df_p[['PROP']]
        # pos_col_prop = str(df['POS_COL'][0])
        pos_row_prop = df_p['POS_ROW'][0]
        
        self.append_df_to_excel(file_to_write, df_prop, sheet_name=group, startrow=pos_row_prop, header=None, index=False, startcol=5)
        logging.debug('\n --->Groups %s<--- \n --->Prop %s<--- \n --->Pos ROW %s<---\n ' % (df_p['GROUPS'][0], df_p['PROP'][0], pos_row_prop))


    # Write Detailed Grouping on Excel
    def writeExcel(self, data, group,file,file_to_write):
        data = data
        df = pds.DataFrame.from_records(data, columns=[ 'TIPO', 'ORIGEM', 'DATA'
                                                      , 'HORA', 'DESTINO', 'CIDADE_DESTINO'
                                                      , 'DURACAO_REAL', 'CUSTO'])
        df['ORIGEM'] = pds.to_numeric(df['ORIGEM'])
        df['DESTINO'] = pds.to_numeric(df['DESTINO'])
        # df['DURACAO_REAL'] = pds.to_datetime(df['DURACAO_REAL']).apply(lambda x: x.strftime(r'%H:%M:%S'))
        # df['DURACAO_REAL'] = pds.to_datetime(df['DURACAO_REAL'], format='%H:%M:%S').dt.time
        # df['DURACAO_REAL'] = pds.to_datetime(df['DURACAO_REAL'], format='%H:%M:%S') + pds.to_timedelta(df['DURACAO_REAL'], unit='s')
        # df['DURACAO_REAL'] = df['DURACAO_REAL'].dt.time
        df['CUSTO'] = pds.to_numeric(df['CUSTO'], downcast='float')
        # print('duracao_real tipo %s ' % (pds.Series(df['DURACAO_REAL']).apply(lambda x: x.strftime(r'%H:%M:%S'))))


        self.append_df_to_excel(file_to_write, df, sheet_name=group, startrow=4, header=None,
                                index=False, startcol=1)
    


    def append_df_to_excel(self, filename, df, sheet_name='Sheet1', startrow=None, startcol=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        # print('filename %s\n df %s\n sheet_name %s\n startrow %s\n startcol %s' % (filename,df,sheet_name,startrow,startcol))
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
        filename : File path or existing ExcelWriter
                    (Example: '/path/to/file.xlsx')
        df : dataframe to save to workbook
        sheet_name : Name of sheet which will contain DataFrame.
                    (default: 'Sheet1')
        startrow : upper left cell row to dump data frame.
                    Per default (startrow=None) calculate the last row
                    in the existing DF and write to the next row...
        truncate_sheet : truncate (remove and recreate) [sheet_name]
                        before writing DataFrame to Excel file
        to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """

        # ignore [engine] parameter if it was passed
        global FileNotFoundError
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pds.ExcelWriter(filename, engine='openpyxl') # pylint: disable=abstract-class-instantiated

        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)
                # date_style = NamedStyle(name='datetime', number_format='h:mm:ss')
                # for cell in sheet_name.cell['H']:
                #     cell.style = date_style
                
                # print("aqui")
                


            # copy existing sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}          
            
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        if startcol is None:
            startcol = 0

        # print('Data Frame %s' % df)
        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, **to_excel_kwargs)

        # workbook = writer.book
        # worksheet = writer.sheets[sheet_name]
        # format1 = workbook.add_format({'num_format':"h:mm:ss"})
        # worksheet.set_column('H:H', None, format1)

        # save the workbook
        writer.save()
   



